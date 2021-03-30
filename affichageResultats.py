# -*- coding: utf-8 -*-
"""
Created on Mon Oct 19 15:33:38 2020

@author: lou
"""
import parametres as p
import random
import openpyxl
from openpyxl.styles import Alignment, Side, Font, Border, PatternFill, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.chart import BarChart,Reference
from openpyxl.chart import PieChart,ProjectedPieChart,Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
import fonctionsMatrices
import inspect
import sys

def creationFichierResultat():
    #nom = input("Nom du fichier:")
    if p.ELECTRICITE_ET_AUTRES and p.INTRANTS_ET_FRET and p.EMBALLAGES_ET_SACHERIE and p.FRET_AVAL:
        nom = "COMPLET "+str(int(100000*random.random()))
    else:
        nom = "test"+str(int(100000*random.random()))
    nom = "Fichier Resultat - "+str(nom)+ " - "+str(p.ANNEE)+".xlsx"
    fichierResultat = openpyxl.Workbook()
    feuille_principale = fichierResultat.active
    feuille_principale.title= "Résultats généraux"  
    
    
    #Tout ce qu'il y a ci après est esthétique
    bd_dotted = Side(style="dotted", color="000000")
    bd_thin = Side(style="thin", color="FF000000")
    
    #Style du fond ffe2aa
    styleFond = NamedStyle(name="Fond")
    styleFond.fill = PatternFill("solid", fgColor= "ffe2aa")
    fichierResultat.add_named_style(styleFond)
    
    #Style du tete de tableau
    styleTeteTab = NamedStyle(name="TeteTab")
    styleTeteTab.border = Border(bottom = bd_dotted, top= bd_dotted)
    styleTeteTab.alignment = Alignment(wrapText = "true",horizontal = "center")
    styleTeteTab.fill = PatternFill("solid", fgColor= "FFFFFF")
    fichierResultat.add_named_style(styleTeteTab)
    
    #Style du valeurs de tableau
    styleVal = NamedStyle(name="Valeurs")
    styleVal.border = Border(left = bd_dotted, right = bd_dotted)
    styleVal.fill = PatternFill("solid", fgColor= "FFF0E1")
    fichierResultat.add_named_style(styleVal)

    #Style des titres
    styleTitre = NamedStyle(name="Titre")
    styleTitre.border = Border(top=bd_dotted, bottom = bd_dotted)
    styleTitre.fill = PatternFill("solid", fgColor= "EBEAF5")
    styleTitre.alignment = Alignment(horizontal = "center", vertical="center")
    styleTitre.font = Font(name = 'Calibri', size =36, italic = False, bold = True, color = 'FF000000')
    fichierResultat.add_named_style(styleTitre)
    
    #Style des entrees de tableau
    styleEntree = NamedStyle(name="Entree")
    styleEntree.border = Border(right = bd_thin)
    styleEntree.fill = PatternFill("solid", fgColor= "FFFFFF")
    fichierResultat.add_named_style(styleEntree)
    
    #Style des tableaux en général
    styleTableau = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    #On applique le fond à toutes les cellules
    for lin in range(1,200):
        for col in [x for x in "ABCDEFGHIJKLMNOPQRSTUVWXYZ"]:
            case = col+str(lin)
            feuille_principale[case].style=styleFond
        for col in [x+y for x in "ABCDEFGHIJKLMNOPQRSTUVWXYZ" for y in "ABCDEFGHIJKLMNOPQRSTUVWXYZ"]:
            case = col+str(lin)
            feuille_principale[case].style=styleFond
    
    #On fusionne les cellules et on écrit le titre
    feuille_principale.merge_cells('B3:Z5')
    feuille_principale["B3"].style = styleTitre
    feuille_principale["B3"] = "Bilan carbone Florentaise - année "+str(p.ANNEE)
    
    #On enregistre les styles en variable globale pour y accéder plus facilement
    p.STYLES_EXCEL = [styleFond, styleTeteTab, styleVal, styleTableau, styleEntree, styleTitre]
    
    #Enfin on essaye de sauvegarder le fichier si l'on peut
    try:
        fichierResultat.save(p.CHEMIN_ECRITURE_RESULTATS + nom)
    except PermissionError:     #Soit à cause du fichier réseau, soit parce que l'excel est ouvert ailleurs
        fonctionsMatrices.print_log_erreur("Permissions insuffisantes pour enregistrer le fichier résultat", inspect.stack()[0][3])
    except FileNotFoundError: #Emplacement inexistant
        fonctionsMatrices.print_log_erreur("Emplacement d'écriture introuvable: "+str(p.CHEMIN_ECRITURE_RESULTATS + nom), inspect.stack()[0][3])
    #Et on renvoie le nom du fichier choisi
    return nom
        
        
        
def sauveFret(sauveFret:list,chemin_ecriture:str, nom:str):
    #On enregistre ici l'onglet log fret qui sert à la calculatrice de produits
    
    #On copie la liste de fret sans les entêtes
    fret_FE = sauveFret[1:]
    
    #On essaye d'ouvrir le fichier
    try:
         log_fret = openpyxl.load_workbook( chemin_ecriture+nom)
    except FileNotFoundError:
         fonctionsMatrices.print_log_erreur("Le document résultat n'est pas trouvé à l'emplacement "+chemin_ecriture+nom, inspect.stack()[0][3])
         sys.exit("Le document résultat n'est pas trouvé à l'emplacement "+ chemin_ecriture+nom)
    
    log_fret = openpyxl.load_workbook( chemin_ecriture+nom)
    #On crée une nouvelle feuille
    feuilleFret = log_fret.create_sheet("log Fret par Matière Première",1)
    #On charge les styles précédemment enregistrés comme variables globales
    styleFond, styleTeteTab, styleVal, styleTableau, styleEntree, styleTitre = p.STYLES_EXCEL
       
    #On applique un style uniforme sur le fond de la feuille
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    for lin in range(1,300):
        for col in [x for x in alphabet]:
            case = col+str(lin)
            feuilleFret[case].style=styleFond
        for col in [x+y for x in alphabet for y in alphabet]:
            case = col+str(lin)
            feuilleFret[case].style=styleFond
    #Redimmensionnement des colonnes pour plus de lisibilité
    tailles = [40,16,12,12,12,14,18,16,18,15,17,17,13,12,16,16]
    for dim in range(len(tailles)):
        feuilleFret.column_dimensions[alphabet[dim]].width = tailles[dim]
        
    feuilleFret.row_dimensions[1].height = 60
    #On bloque la vue de la première ligne et colonne
    feuilleFret.freeze_panes = "B2"
    
    #On note ici le nom de la colonne et le numéro de la grandeur dans une ligne de fret_FE
    colonne_et_correspondance = [[sauveFret[0][i] , i] for i in range(len(sauveFret[0]))]
    #Cela permet de reprendre les entêtes du tableau telles quelles.    
    
    #Inscrit les entêtes du tableau (et applique le style adéquat)
    for i in range(1,len(colonne_et_correspondance)+1):
        feuilleFret.cell(row=1, column=i).value = colonne_et_correspondance[i-1][0]
        feuilleFret.cell(row=1, column=i).style = styleTeteTab
        feuilleFret.cell(row=1, column=i).alignment = Alignment(wrapText=True, horizontal='center', vertical = 'center')
        
    #Inscrit les valeurs dans le tableau ligne par ligne
    for i in range(1,len(fret_FE)):
        # print(len(fret_FE[i]))
        for j in range(len(colonne_et_correspondance)):
            # print("i:"+str(i)+", j:"+str(j))

            feuilleFret.cell(row=i+1, column=j+1).value = fret_FE[i][colonne_et_correspondance[j][1]]
            #On applique un style différent à la première colonne
            if j==0:
                feuilleFret.cell(row=i+1, column=1).style = styleEntree
            else:
                feuilleFret.cell(row=i+1, column=j+1).style = styleVal
                if j in [1,2,13,14,15]:
                    feuilleFret.cell(row=i+1, column=j+1).number_format = '0'
                elif j in [3,4,5,6,7,8,9,10,11,12]:
                    feuilleFret.cell(row=i+1, column=j+1).number_format = '0.0'
                    
    emphase_colonne_n(feuilleFret, 13)
    emphase_colonne_n(feuilleFret, 14)
    #Affichage de graphe
    if True:
        #On crée un graphique en barre
        graphique = BarChart()
        #Style horizontal
        graphique.type = "bar"
        graphique.style = 12 #whatever
        graphique.title = "FE des matières premières, fabrication + fret moyen "+str(p.ANNEE)+" (tCO2e/t)"
        graphique.grouping = "stacked"
        #100% de recouvrement pour avoir la seconde catégorie (axe2) par dessus
        graphique.overlap = 100 
        #On prend en abscisses la première colonne
        mp = Reference(feuilleFret, min_col = 1, min_row = 2, max_row = len(fret_FE))
        #Les ordonnées sont en colonnes 9 10 12
        for i in [9,10,12]:
            #On ajoute une à une les séries
            valeurs =  Reference(feuilleFret, min_col=i, min_row=1, max_row=len(fret_FE))
            graphique.add_data(valeurs, titles_from_data=True)
        graphique.set_categories(mp)
        
        #Dimensions du graphique
        graphique.height = 103
        graphique.width = 20
        graphique.x_axis.scaling.orientation = "maxMin"
        cp = CharacterProperties(sz=800) #Taille du label entre 100 et 400 000
        graphique.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
        
        #Création du second axe de BC total
        axe2 = BarChart() #idem
        axe2.type = 'bar'
        val2 = Reference(feuilleFret, min_col=15, min_row=1, max_row=len(fret_FE))
        axe2.set_categories(mp)
        axe2.add_data(val2, titles_from_data=True, from_rows=False)
        serie = axe2.series[0]
        #On rend cette série transparente avec les bords rouges
        serie.graphicalProperties.line.solidFill = "FF0000"
        serie.graphicalProperties.noFill = True
        axe2.overlap = -50 #{ça fonctionne comme ça}
        axe2.y_axis.axId = 200
        axe2.y_axis.title = "Bilan Carbone total de la MP"
        axe2.y_axis.crosses = "min"
        
        #On ahoute l'axe finalement
        graphique += axe2
        
        graphique.legend.legendPos = 't'
        graphique.legend.tagname= 'Légende'    
        graphique.y_axis.title = 'Emissions Carbones (kgCO2e/t)'
        feuilleFret.add_chart(graphique, "C5")        
        
    #Si on a la possibilité, on sauvegarde
    try:
        log_fret.save(chemin_ecriture+nom)
    except PermissionError:     #Soit à cause du fichier réseau, soit parce que l'excel est ouvert ailleurs
        fonctionsMatrices.print_log_erreur("Permissions insuffisantes pour enregistrer le fichier résultat", inspect.stack()[0][3])
    except FileNotFoundError: #Emplacement inexistant
        fonctionsMatrices.print_log_erreur("Emplacement d'écriture introuvable: "+str(p.CHEMIN_ECRITURE_RESULTATS + nom), inspect.stack()[0][3])
    
            
def sauve_resultats_generaux(nom_fichier:str):
    usines = []
    for x in p.CP_SITES_FLORENTAISE:
        if x[3]<=p.ANNEE:
            usines.append(x)
    entetes_resultats_generaux(usines, nom_fichier)
    return None
    
def entetes_resultats_generaux(usines:list, nom_fichier:str):
    styleFond, styleTeteTab, styleVal, styleTableau, styleEntree, styleTitre = p.STYLES_EXCEL
    
    ligne_depart = 10
    try:
         document = openpyxl.load_workbook(p.CHEMIN_ECRITURE_RESULTATS+nom_fichier)
    except FileNotFoundError:
         fonctionsMatrices.print_log_erreur("Le document résultat n'est pas trouvé à l'emplacement "+p.CHEMIN_ECRITURE_RESULTATS+nom_fichier, inspect.stack()[0][3])
         sys.exit("Le document résultat n'est pas trouvé à l'emplacement "+ p.CHEMIN_ECRITURE_RESULTATS+nom_fichier)
         
    feuilleprinc = document.get_sheet_by_name(document.get_sheet_names()[0])
    feuilleprinc.column_dimensions['B'].width = 24
    feuilleprinc.column_dimensions['C'].width = 0
    feuilleprinc.column_dimensions['D'].width = 20
    feuilleprinc.column_dimensions['E'].width = 10
    
    FretAmont = ["MP bateau", "MP route"]
    FretAval = ["Ventes pro","Ventes Jardineries", "Interdepot"]
    Intrants = ["Matières Premières", "CO2 issu de la tourbe"]
    HorsEnergie = ["Protoxyde d'azote"]
    Energie = ["Electricité", "Fuel"]
    Emballages = ["PE neuf", "PE recyclé", "PVC", "Carton"]

    entetes = FretAmont+FretAval+Intrants+HorsEnergie+Energie+Emballages
    
    colonnes = ["Poste d'émission", "Site", "Empreinte carbone (tCO2e)","", "Emission par m3 (kgCO2e/m3)", ""]
    
    for i in range(len(colonnes)):
        feuilleprinc.cell(row = ligne_depart, column = i+2).value = colonnes[i]
        feuilleprinc.cell(row=ligne_depart, column=i+2).style = styleTeteTab
    i = 0
    for poste in entetes:
        for j in range(len(usines)):
            feuilleprinc["B"+str(i+ligne_depart+1)] = poste
            feuilleprinc["C"+str(i+ligne_depart+1)] = str(usines[j][0]) + " " + usines[j][1]
            i+=1
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    tableau = Table(displayName = "Resultats_principaux", ref = "B"+str(ligne_depart)+":"+alphabet[len(entetes)]+str(ligne_depart+i))  
    tableau.tableStyleInfo = styleTableau
    production_totale = sum([x[1]for x in p.PROD_PAR_SITE])
    feuilleprinc['C7'] = "Total:"
    feuilleprinc['D7'] = "=TEXT(SUM(D11:D"+str(199)+"), \"# ##0\")&+\" tCO2e\""
    feuilleprinc['E7'] = "et"
    feuilleprinc['F7'] = "=TEXT(1000*SUM(D11:D"+str(199)+")/"+str(production_totale)+", \"# ##0\")&+\" kgCO2e/m3\""
    rouge = openpyxl.styles.colors.Color(rgb='00FF0000')
    for c in "CDEF":
        feuilleprinc[c+"7"].font = Font(size=18, color = rouge, bold=True)
    #Si on a la possibilité, on sauvegarde
    try:
        document.save(p.CHEMIN_ECRITURE_RESULTATS+nom_fichier)
    except PermissionError:     #Soit à cause du fichier réseau, soit parce que l'excel est ouvert ailleurs
        fonctionsMatrices.print_log_erreur("Permissions insuffisantes pour enregistrer le fichier résultat", inspect.stack()[0][3])
    except FileNotFoundError: #Emplacement inexistant
        fonctionsMatrices.print_log_erreur("Emplacement d'écriture introuvable: "+str(p.CHEMIN_ECRITURE_RESULTATS + nom_fichier), inspect.stack()[0][3])
    

#Edite le style de toute une colonne pour la mettre en avant (ici en gras)
def emphase_colonne_n(feuille, n:int):
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ" #Flemme de faire fonctionner pour n>26... (oupsie)
    for i in range(1,feuille.max_row-1):
        feuille[alphabet[n]+str(i)].font = Font(bold=True)
    return True

def sauveTout(nom:str, recap:list):
    try:
         log_fret = openpyxl.load_workbook(p.CHEMIN_ECRITURE_RESULTATS+nom)
    except FileNotFoundError:
         fonctionsMatrices.print_log_erreur("Le document résultat n'est pas trouvé à l'emplacement "+p.CHEMIN_ECRITURE_RESULTATS+nom, inspect.stack()[0][3])
         sys.exit("Le document résultat n'est pas trouvé à l'emplacement "+ p.CHEMIN_ECRITURE_RESULTATS+nom)
    feuille = log_fret.get_sheet_by_name(log_fret.get_sheet_names()[0])
    styleFond, styleTeteTab, styleVal, styleTableau, styleEntree, styleTitre = p.STYLES_EXCEL
    
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    tableau = [[u[0], u[1][1], u[2][0], u[2][1], u[2][2], u[2][3]]for u in p.MATRICE_RESULTAT]
    
    for li in range(1,len(tableau[1:])+1):
        for col in range(len(tableau[li])):
            if col>0:
                feuille.cell(row = li+10, column = col+2).value=tableau[li][col]
            else:
                feuille.cell(row = li+10, column = col+2).value=tableau[li][col]+" - "+tableau[li][col+1]
            if col<=1:
                feuille.cell(row=li+10, column=col+2).style = styleEntree
            else:
                feuille.cell(row=li+10, column=col+2).style = styleVal
                if col in [2,4]:
                    feuille.cell(row = li+10, column = col+2).number_format = '0.0'
        
    tab1 = Table(displayName="Tableau_general", ref="A2:F"+str(len(p.MATRICE_RESULTAT)))
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab1.tableStyleInfo = style
    
    # rule = DataBarRule(start_type='num', start_value=0, end_type='percentile', end_value='100',color="FF638EC6", showValue="None", minLength=None, maxLength=None)
    # feuille.conditional_formatting.add("C2:C"+str(len(p.MATRICE_RESULTAT)), rule)
    feuille.column_dimensions['A'].width = 10
    feuille.column_dimensions['B'].width = 17
    feuille.column_dimensions['C'].width = 0
    feuille.column_dimensions['D'].width = 25
    feuille.column_dimensions['E'].width = 6
    feuille.column_dimensions['F'].width = 25
    feuille.column_dimensions['G'].width = 10
    

     #Affichage du premier graphe général (avec tous les postes)
    if True:
        graphique = BarChart()
        graphique.type = "bar"
        graphique.style = 12
        graphique.title = "Bilan carbone total par poste et par site"
        graphique.grouping = "standard"
        graphique.overlap = 00
        
        position_row = 9
        position_col = 1
        postes = Reference(feuille, min_col = position_col+1, min_row = position_row+2, max_row = position_row+len(p.MATRICE_RESULTAT)-1)
        for i in [3]:
            valeurs =  Reference(feuille, min_col=position_col+i, min_row=position_row+1, max_row=position_row+len(p.MATRICE_RESULTAT)-1)
            graphique.add_data(valeurs, titles_from_data=True)
        graphique.set_categories(postes)
        graphique.height = 103
        graphique.width = 20
        secaxe = BarChart()
        valsecond = Reference(feuille, min_col=position_col+i+2, min_row=position_row+1, max_row=position_row+len(p.MATRICE_RESULTAT)-1)
        secaxe.add_data(valsecond, titles_from_data=True)
        secaxe.y_axis.axId = 200
        graphique.x_axis.scaling.orientation = "maxMin"
        secaxe.type = "bar"
        secaxe.y_axis.title = "Ramenées à la production (kgeCO2/m3)"
        cp = CharacterProperties(sz=800) #Taille du label entre 100 et 400 000
        graphique.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
        graphique.legend.legendPos = 't'
        graphique.legend.tagname= 'Légende'    
        graphique.y_axis.title = 'Emissions Carbones (teCO2)'
        graphique.y_axis.crosses = "min"
        serie = secaxe.series[0]
        serie.graphicalProperties.line.solidFill = "FF0000"
        serie.graphicalProperties.noFill = True
        graphique += secaxe
        feuille.add_chart(graphique, "I10")
    
    
    matRes_inv_tot,matRes_inv_m3  = renversermatres(p.MATRICE_RESULTAT)
    nb_postes = len(matRes_inv_tot)-1
    nb_sites = len(matRes_inv_tot[0])-1
    
    
    
    
    # Changement de feuille du classeur
    feuille = log_fret.create_sheet("Resultats detailles",1)
    for lin in range(1,200):
        for col in [x for x in "ABCDEFGHIJKLMNOPQRSTUVWXYZ"]:
            case = col+str(lin)
            feuille[case].style=styleFond
        for col in [x+y for x in "ABCDEFGHIJKLMNOPQRSTUVWXYZ" for y in "ABCDEFGHIJKLMNOPQRSTUVWXYZ"]:
            case = col+str(lin)
            feuille[case].style=styleFond
    feuille.column_dimensions['A'].width = 10
    feuille.column_dimensions['B'].width = 31
    for c in "CDEFGHIJKL":
        feuille.column_dimensions[c].width = 11

    
    feuille.merge_cells('B3:Z5')
    for c in "BCDEFGHIJKLMNOPQRSTUVWXYZ":
        feuille[c+"3"].style = styleTitre
    feuille["B3"] = "Résultats par poste et par site - année "+str(p.ANNEE)
    
    ligne = 9
    colonne = 1
    feuille.cell(row=ligne, column=1+colonne).value ="Emissions totales, par site et par poste, en tCO2e"
    feuille.merge_cells('B'+str(ligne)+':L'+str(ligne))
    feuille.cell(row=ligne, column=1+colonne).style = styleEntree
    feuille.cell(row=ligne, column=1+colonne).font = Font(bold=True)
    ligne+=2
    for l in range(len(matRes_inv_tot)):
        for head in range(len(matRes_inv_tot[0])):
            feuille.cell(row=ligne, column=head+1+colonne).value =matRes_inv_tot[l][head]
            if l == 0:
                feuille.cell(row=ligne, column=head+1+colonne).style = styleTeteTab                
            elif head<1:
                feuille.cell(row=ligne, column=head+1+colonne).style = styleEntree
            else:
                feuille.cell(row=ligne, column=head+1+colonne).style = styleVal
                feuille.cell(row = ligne, column = head+1+colonne).number_format = '0'
        ligne +=1
    ligne +=1
    feuille.cell(row=ligne, column=1+colonne).value ="Emissions totales ramenées à la production, par site et par poste, en kgCO2e/m3"
    feuille.merge_cells('B'+str(ligne)+':L'+str(ligne))
    feuille.cell(row=ligne, column=1+colonne).style = styleEntree
    feuille.cell(row=ligne, column=1+colonne).font = Font(bold=True)
    ligne +=2
    for l in range(len(matRes_inv_m3)):
        for head in range(len(matRes_inv_m3[0])):
            feuille.cell(row=ligne, column=head+1+colonne).value =matRes_inv_m3[l][head]
            if l == 0:
                feuille.cell(row=ligne, column=head+1+colonne).style = styleTeteTab                
            elif head<1:
                feuille.cell(row=ligne, column=head+1+colonne).style = styleEntree
            else:
                feuille.cell(row=ligne, column=head+1+colonne).style = styleVal
                feuille.cell(row = ligne, column = head+1+colonne).number_format = '0.0'
        ligne +=1
    ligne += 2
    sites = Reference(feuille, min_row = 11, min_col = 2+colonne, max_col = nb_sites+1+colonne)
    graphique1 = BarChart()
    graphique1.title = "Emissions totales des sites par poste"
    graphique1.type = "bar";graphique1.style= 10;graphique1.grouping = "stacked";graphique1.overlap = 100
    graphique1.height = 15;graphique1.width = 25
    graphique1.x_axis.scaling.orientation = "maxMin"
    cp = CharacterProperties(sz=1000) #Taille du label entre 100 et 400 000
    graphique1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    graphique1.varyColors=True
    valeurs1 =  Reference(feuille, min_row=12, min_col=1+colonne, max_col=nb_sites+1+colonne, max_row = nb_postes+11)

    graphique1.add_data(valeurs1, titles_from_data=True, from_rows=True)
    graphique1.set_categories(sites)
    graphique1.y_axis.title = 'Emissions Carbones (tCO2e)'
    feuille.add_chart(graphique1, "O7")
    
    
    offsetg2 = nb_postes+4
    graphique2 = BarChart()
    graphique2.title = "Emissions totales des sites par poste"
    graphique2.type = "bar";graphique2.style= 10;graphique2.grouping = "stacked";graphique2.overlap = 100
    graphique2.height = 15;graphique2.width = 25
    graphique2.x_axis.scaling.orientation = "maxMin"
    cp = CharacterProperties(sz=1000) #Taille du label entre 100 et 400 000
    graphique2.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    graphique2.varyColors=True
    valeurs2 =  Reference(feuille, min_row=12+offsetg2, min_col=1+colonne, max_col=nb_sites+1+colonne, max_row = nb_postes+11+offsetg2)

    graphique2.add_data(valeurs2, titles_from_data=True, from_rows=True)
    graphique2.set_categories(sites)
    graphique2.y_axis.title = 'Emissions Carbone par volume vendu (kgCO2e/m3)'
    feuille.add_chart(graphique2, "O38")
        
    if p.ELECTRICITE_ET_AUTRES and p.INTRANTS_ET_FRET and p.EMBALLAGES_ET_SACHERIE and p.FRET_AVAL:
        ligne = enregistre_resultat_resume(recap, ligne, colonne, feuille, nb_postes)
    
    
    l = 1
    while l<25 and feuille.cell(row=l, column=2).value!= 'MP bateau':
        l+=1
    l0 =l
    if l!=25:
        feuille.cell(row=l-1, column=3+nb_sites).value = "Total"
        feuille.cell(row=l-1, column=3+nb_sites).font = Font(bold=True)
        while feuille.cell(row=l, column=2).value!= None and l<50:
            feuille.cell(row=l, column=3+nb_sites).value = "=SUM(C"+str(l)+":"+alphabet[nb_sites+1]+str(l)+")"
            feuille.cell(row=l, column=3+nb_sites).font = Font(italic=True)
            feuille.cell(row=l, column=3+nb_sites).number_format = '0.0'
            l+=1
        feuille.cell(row=l, column=3+nb_sites).value = "=SUM("+alphabet[nb_sites+2]+str(l0)+":"+alphabet[nb_sites+2]+str(l-1)+")"
        feuille.cell(row=l, column=3+nb_sites).font = Font(bold=True)
        feuille.cell(row=l, column=3+nb_sites).number_format = '0.0'
    
    
    
    #Si on a la possibilité, on sauvegarde
    try:
        log_fret.save(p.CHEMIN_ECRITURE_RESULTATS+nom)
    except PermissionError:     #Soit à cause du fichier réseau, soit parce que l'excel est ouvert ailleurs
        fonctionsMatrices.print_log_erreur("Permissions insuffisantes pour enregistrer le fichier résultat", inspect.stack()[0][3])
    except FileNotFoundError: #Emplacement inexistant
        fonctionsMatrices.print_log_erreur("Emplacement d'écriture introuvable: "+str(p.CHEMIN_ECRITURE_RESULTATS + nom), inspect.stack()[0][3])
    
    return None



def enregistre_resultat_resume(recap, ligne, colonne, feuille, nb_postes):
    styleFond, styleTeteTab, styleVal, styleTableau, styleEntree, styleTitre = p.STYLES_EXCEL
    feuille.cell(row=ligne, column=1+colonne).value ="Regroupement de résultats, les postes principaux"
    feuille.merge_cells('B'+str(ligne)+':L'+str(ligne))
    feuille.cell(row=ligne, column=1+colonne).style = styleEntree
    feuille.cell(row=ligne, column=1+colonne).font = Font(bold=True)
    ligne += 2
    
    for li in range(len(recap)):
        for co in range(len(recap[0])):
            feuille.cell(row=ligne, column=1+colonne+co).value = recap[li][co]
            if li == 0:
                feuille.cell(row=ligne, column=1+colonne+co).style = styleTeteTab                
            elif co<1:
                feuille.cell(row=ligne, column=1+colonne+co).style = styleEntree
            else:
                feuille.cell(row=ligne, column=1+colonne+co).style = styleVal
                feuille.cell(row = ligne, column = 1+colonne+co).number_format = '0.0'
        ligne+=1
        
    feuille.cell(row=ligne, column=2).value = "EGES tot dues à la tourbe"
    for c in range(3, 11):
        feuille.cell(row=ligne, column=c).style = styleVal
        feuille.cell(row=ligne, column=8).number_format = '#.0'
    feuille.cell(row=ligne, column=2).style = styleEntree
    feuille.cell(row=ligne, column=3).value = p.COMPARAISON_TOURBE[0]
    feuille.cell(row=ligne, column=3).number_format = '#.0'
    feuille.cell(row=ligne, column=6).value = p.COMPARAISON_TOURBE[1]
    feuille.cell(row=ligne, column=6).number_format = '#.0'
    feuille.cell(row=ligne, column=8).value = p.COMPARAISON_TOURBE[2]
    feuille.cell(row=ligne, column=8).number_format = '#.0'
    feuille.cell(row=ligne, column=10).value = "=SUM(C"+str(ligne)+":I"+str(ligne)+")"
    feuille.cell(row=ligne, column=10).number_format = '#.0'
    ligne +=1
    
    for li in range(1,len(recap)):
        for co in range(len(recap[0])):
            if co<1:
                feuille.cell(row=ligne, column=1+colonne+co).value = recap[li][co]+ " (%)"
                feuille.cell(row=ligne, column=1+colonne+co).style = styleEntree
            elif co <=len(recap[0])-1 and li ==1:
                feuille.cell(row=ligne, column=1+colonne+co).value = recap[li][co]/recap[1][-1]
                feuille.cell(row=ligne, column=1+colonne+co).style = styleVal
                feuille.cell(row=ligne, column=1+colonne+co).number_format = '0.0%'
            elif co <=len(recap[0])-1 and li ==2:
                feuille.cell(row=ligne, column=1+colonne+co).value = recap[li][co]/recap[2][-1]
                feuille.cell(row=ligne, column=1+colonne+co).style = styleVal
                feuille.cell(row=ligne, column=1+colonne+co).number_format = '0.0%'
            
        ligne+=1
    
    
    
    offset3 = 20+(2*nb_postes)
    sites = Reference(feuille, min_row = offset3, min_col = 2+colonne, max_col = len(recap[0])+colonne-1)
    graphique3 = BarChart()
    graphique3.title = "Emissions totales des sites par poste"
    graphique3.type = "col";graphique3.style= 10;graphique3.grouping='standard';graphique3.overlap = 25
    graphique3.height = 12;graphique3.width = 25
    graphique3.x_axis.scaling.orientation = "maxMin"
    cp = CharacterProperties(sz=1000) #Taille du label entre 100 et 400 000
    graphique3.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    graphique3.varyColors=True
    valeurs3 =  Reference(feuille, min_row=1+offset3, min_col=1+colonne, 
                          max_col=len(recap[0])+colonne-1, max_row =1+offset3 )

    
    graphique3.y_axis.crosses = "max"
    graphique3.add_data(valeurs3, titles_from_data=True, from_rows=True)
    graphique3.set_categories(sites)
    graphique3.y_axis.title = 'Emissions Carbones globales (tCO2e)'
    # graphique3.layout=Layout(manualLayout=ManualLayout(x=0.1, y=0.050,h=0.80, w=0.90,))
    graphique3.legend.layout = Layout(manualLayout=ManualLayout(yMode='edge',xMode='edge',x=0, y=0.9,h=0.1, w=0.5))
    feuille.add_chart(graphique3, "B"+str(offset3+7))
    
    tarte = PieChart()
    labels = Reference(feuille, min_row = offset3, min_col = 2+colonne, max_col = len(recap[0])+colonne-1)
    for i in [0]:
        data = Reference(feuille, min_row=offset3+1+i, min_col=1+colonne, max_col=len(recap[0])+colonne-1)
        tarte.add_data(data, titles_from_data=True, from_rows=True) 
    tarte.set_categories(labels)
    tarte.dataLabels = DataLabelList() 
    tarte.dataLabels.position ='bestFit'
    tarte.dataLabels.showCat  = True
    tarte.dataLabels.showPercent = True
    tarte.dataLabels.separator = ":"
    tarte.height = 8.15;tarte.width = 9
    tarte.title = "Part des postes dans les émissions totales"
    set_chart_title_size(tarte, size=1200)
    feuille.add_chart(tarte, "M"+str(offset3+16))
    
    
    
    
    
    return ligne
    
    
    
        
        
def ecrire_messages_derreur(nom:str):
    #Tous les messages d'erreur
    try:
         log_erreurs = openpyxl.load_workbook( p.CHEMIN_ECRITURE_RESULTATS+nom)
    except FileNotFoundError:
         fonctionsMatrices.print_log_erreur("Le document résultat n'est pas trouvé à l'emplacement "+ p.CHEMIN_ECRITURE_RESULTATS+nom, inspect.stack()[0][3])
         sys.exit("Le document résultat n'est pas trouvé à l'emplacement "+  p.CHEMIN_ECRITURE_RESULTATS+nom)
    feuilleErreurs = log_erreurs.create_sheet("Erreurs et informations",-1)
    styleFond, styleTeteTab, styleVal, styleTableau, styleEntree, styleTitre = p.STYLES_EXCEL
    
    if len(p.MESSAGES_DERREUR)>1:
        for message in p.MESSAGES_DERREUR:
            feuilleErreurs.append(message)
    feuilleErreurs.column_dimensions['A'].width = 9
    feuilleErreurs.column_dimensions['B'].width = 160
    feuilleErreurs.column_dimensions['C'].width = 28
    
    
    #Les codes postaux erronés
    if p.EMBALLAGES_ET_SACHERIE or p.INTRANTS_ET_FRET or p.FRET_AVAL:
        if len(p.LISTE_CODES_POSTAUX_ERRONES)>1:
            feuilleErreursCP = log_erreurs.create_sheet("Codes postaux à corriger",-1)
            feuilleErreursCP.column_dimensions['A'].width = 14
            feuilleErreursCP.column_dimensions['B'].width = 14
            for cp in p.LISTE_CODES_POSTAUX_ERRONES:
                feuilleErreursCP.append(cp)
    
    
    #Les assimilations de matières premières
    if p.INTRANTS_ET_FRET:
        if len(p.LISTE_ASSIMILATIONS)>1:
            feuilleAssim = log_erreurs.create_sheet("Assimilations de MP",-1)
            feuilleAssim.column_dimensions['A'].width = 40
            feuilleAssim.column_dimensions['B'].width = 40
            for message in p.LISTE_ASSIMILATIONS:
                feuilleAssim.append(message)
    
   
     #Si on a la possibilité, on sauvegarde
    try:
        log_erreurs.save(p.CHEMIN_ECRITURE_RESULTATS+nom)
    except PermissionError:     #Soit à cause du fichier réseau, soit parce que l'excel est ouvert ailleurs
        fonctionsMatrices.print_log_erreur("Permissions insuffisantes pour enregistrer le fichier résultat", inspect.stack()[0][3])
    except FileNotFoundError: #Emplacement inexistant
        fonctionsMatrices.print_log_erreur("Emplacement d'écriture introuvable: "+str(p.CHEMIN_ECRITURE_RESULTATS + nom), inspect.stack()[0][3])
    return None
        
        
        
def renversermatres(matRes_h:list):
    matRes = matRes_h[1:].copy()
    #D'abord on récupère la liste des sites
    elem0 = matRes[0]
    sites = [k[1][1] for k in matRes if k[0]==elem0[0]]
    res = [["Poste d'émission"]+sites]
    postes = fonctionsMatrices.liste_unique((fonctionsMatrices.extraire_colonne_n(0, matRes)))
    for po in postes:
        res.append([po]+[0 for x in sites])
    res_parM3= [["Poste d'émission"]+sites]
    for po in postes:
        res_parM3.append([po]+[0 for x in sites])
    for po in range(1,len(res)):
        poste = res[po][0]
        for s in range(1,len(res[0])):
            site = res[0][s]
            for p2 in range(len(matRes)):
                if matRes[p2][0]==poste and matRes[p2][1][1]==site:
                    res[po][s] = matRes[p2][2][0]
                    res_parM3[po][s] = matRes[p2][2][2]
                    break
    return res,res_parM3





def set_chart_title_size(chart, size=1400):
    paraprops = ParagraphProperties()
    paraprops.defRPr = CharacterProperties(sz=size)

    for para in chart.title.tx.rich.paragraphs:
        para.pPr=paraprops 

